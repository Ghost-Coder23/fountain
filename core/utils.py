from django.core.exceptions import PermissionDenied
from functools import wraps
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.urls import reverse
from schools.models import SchoolUser, School
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import csv

def get_default_school():
    """Get the first (and only) school in single-tenancy mode"""
    return School.objects.first()

def send_welcome_email(request, user, school, is_new_user=True):
    """
    Sends a welcome email to a new or existing user added to a school.
    Includes a password reset link if it's a new user.
    """
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    
    # Use request.build_absolute_uri to ensure the link points to the current domain
    reset_url = request.build_absolute_uri(
        reverse('accounts:password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
    )
    
    subject = f'Welcome to {school.name}'
    message = render_to_string('accounts/welcome_email.txt', {
        'user': user,
        'school': school,
        'reset_link': reset_url,
        'is_new_user': is_new_user,
    })
    
    send_mail(
        subject,
        message,
        'noreply@educore.com',
        [user.email],
        html_message=message
    )

def school_role_required(roles):
    """
    Decorator for views that checks if the user has one of the required roles 
    in the current school context (single-tenancy mode).
    """
    if isinstance(roles, str):
        roles = [roles]

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return PermissionDenied("Authentication required")
            
            school = get_default_school()
            if not school:
                return PermissionDenied("No school found in the system")

            try:
                membership = SchoolUser.objects.get(
                    user=request.user, 
                    school=school, 
                    is_active=True
                )
                if membership.role in roles or request.user.is_superuser:
                    return view_func(request, *args, **kwargs)
                else:
                    raise PermissionDenied("You do not have the required role")
            except SchoolUser.DoesNotExist:
                raise PermissionDenied("You are not a member of this school")

        return _wrapped_view
    return decorator

class SchoolRoleMixin:
    """Mixin for class-based views to enforce school role requirements (single-tenancy mode)."""
    required_roles = []

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            raise PermissionDenied("Authentication required")

        school = get_default_school()
        if not school:
            raise PermissionDenied("No school found in the system")

        try:
            membership = SchoolUser.objects.get(
                user=request.user, 
                school=school, 
                is_active=True
            )
            if membership.role in self.required_roles or request.user.is_superuser:
                return super().dispatch(request, *args, **kwargs)
            else:
                raise PermissionDenied("You do not have the required role")
        except SchoolUser.DoesNotExist:
            raise PermissionDenied("You are not a member of this school")

def export_to_excel(data, headers, filename="export.xlsx"):
    """
    Export data to Excel file.
    
    Args:
        data: List of lists containing the data rows
        headers: List of column headers
        filename: Name of the Excel file to download
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].bestFit = True
        ws.column_dimensions[get_column_letter(col_num)].auto_size = True
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save workbook to response
    wb.save(response)
    return response

def export_to_csv(data, headers, filename="export.csv"):
    """
    Export data to CSV file.
    
    Args:
        data: List of lists containing the data rows
        headers: List of column headers
        filename: Name of the CSV file to download
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(data)
    
    return response
